import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# wb = xl.load_workbook('transactions.xlsx') # workbook
# sheet = wb['Sheet1']
# cell = sheet['a1']   # sheet.cell(1,1) - both are same
# cell = sheet.cell(1, 1)
# # print(sheet.max_row)
#
# # for row in range(1, sheet.max_row + 1): # it will generate 1,2,3 but not 4
# for row in range(2, sheet.max_row + 1):
#     # print(row)
#     cell = sheet.cell(row, 3)
#     # print(cell.value)
#     corrected_price = cell.value * 0.9
#     corrected_price_cell = sheet.cell(row, 4)
#     corrected_price_cell.value = corrected_price
def process_workbook(filename):
    wb = xl.load_workbook(filename)  # workbook
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
    # print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)
